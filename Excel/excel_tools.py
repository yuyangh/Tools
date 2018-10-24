# -*- coding: utf-8 -*-

# read and write 2007 excel
import numbers
import openpyxl
import json
import os


def read_07_Excel_sheet(path, sheet_name='Sheet1'):
    wb = openpyxl.load_workbook(path)
    sheet = wb[sheet_name]
    return sheet


def print_entire_sheet(sheet):
    '''

    :param sheet: a sheet object
    :return:
    '''
    for row in sheet.rows:
        for cell in row:
            print(cell.value, "\t", end="")
        print()


def write_to_json(file_path, data):
    with open(file_path, 'w') as outfile:
        json.dump(data, outfile)


def getAllNameJSONDataURL(path):
    '''
    @:return a list of map {name:string,url:string,JSONdata:json} from Excel document
    '''
    data = []
    wb = openpyxl.load_workbook(path)
    sheet = wb['Sheet1']
    for row in sheet.rows:
        if (len(row)) >= 8:
            name = row[2].value
            url = row[7].value
            JSONdata = json.loads(row[8].value)
            data.append({"name": name, 'url': url, 'JSONdata': JSONdata})
    return data


def getNameURLList(path):
    '''

    :param path: excel document path
    :return: a list of map {name:string, url:string, JSONdata:json}
    '''
    data = []
    wb = openpyxl.load_workbook(path)
    sheet = wb['图片识别情况']
    for row in sheet.rows:
        if (len(row)) >= 3:
            name = row[2].value
            url = row[0].value
            JSONdata = ""
            data.append({"name": name, 'url': url, 'JSONdata': JSONdata})
    return data


def excel07_fill_in_blanks(path, row=None, col=None, range=[0, 0]):
    wb = openpyxl.load_workbook(path)
    sheet = wb['Sheet1']
    previous_value = None
    if (col != None):
        for cell in sheet[col][range[0]:range[1]]:
            if (cell.value != None):
                print(cell.value)
                previous_value = cell.value
            else:
                print(previous_value)
    if (row != None):
        for cell in sheet[row][range[0]:range[1]]:
            if (cell.value != None):
                print(cell.value)
                previous_value = cell.value
            else:
                print(previous_value)


def get_selected_mentors(sheet, mentee_info_dict):
    '''

    :param sheet:
    :param mentee_info_dict: key : mentee email, value: a list has email, linkedIn, facebook
    :return: mentor_mentee_dict_1st key : mentor index, value: list of mentee emails
    '''
    mentor_mentee_dict_1st = dict()
    for row in sheet.rows:
        if row[2].value.lower() == "mentee":
            email = row[1].value.strip().lower()
            # wills are index of each mentor
            first_will = int(row[4].value.split(" ")[0])
            second_will = int(row[5].value.split(" ")[0])

            # a dictionary: key : mentor index, value: list of mentee emails
            if first_will not in mentor_mentee_dict_1st:
                mentor_mentee_dict_1st[first_will] = [email]
            else:
                mentor_mentee_dict_1st[first_will].append(email)

            # todo add second will
            if second_will not in mentor_mentee_dict_1st:
                mentor_mentee_dict_1st[second_will] = [email]
            else:
                mentor_mentee_dict_1st[second_will].append(email)
            # append linkedIn, facebook
            mentee_info_dict[email].append(row[6].value)
            mentee_info_dict[email].append(row[7].value)

    # for key in mentor_mentee_dict_1st.keys():
    #     print("{:3}".format(key),end=" ")
    #     # "{:3} mentees".format(len(mentor_mentee_dict_1st[key]))
    #     for email in mentor_mentee_dict_1st[key]:
    #         print(email,end=" ")
    #     print()
    return mentor_mentee_dict_1st


def get_mentee_info(sheet):
    '''

    :param sheet:
    :return: mentee_info_dict, key: email, value: name
    '''
    mentee_info_dict = dict()
    for row in sheet.rows:
        if isinstance(row[0].value, int):
            index = int(row[0].value)
            name = row[1].value.strip()
            email = row[2].value.strip().lower()
            # print(index,name,email)
            mentee_info_dict[email] = [name]
    print(mentee_info_dict)
    return mentee_info_dict


def get_mentor_info(sheet):
    '''
    read mentor index, name, email
    :param sheet:
    :return: mentor_info_dict: key: index, values: name and email
    '''
    mentor_info_dict = dict()
    for row in sheet.rows:
        if isinstance(row[0].value, numbers.Rational):
            index = int(row[0].value)
            name = row[1].value.strip()
            email = row[2].value.strip().lower()
            # print(index,name,email)
            mentor_info_dict[index] = name + " " + email
    print(mentor_info_dict)
    return mentor_info_dict


def get_mentor_info_json(sheet):
    '''
    read mentor index, name, email
    :param sheet:
    :return: mentor_info_dict: key: email, values: name and email
    '''
    mentor_info_dict = dict()
    for row in sheet.rows:
        if isinstance(row[0].value, numbers.Rational):
            index = int(row[0].value)
            name = row[1].value.strip()
            email = row[2].value.strip().lower()
            # print(index,name,email)
            if email not in mentee_info_dict.keys():
                mentor_info_dict[email] = {'index': index, 'name': name}
            else:
                print("DUPLICATE EMAIL")
    print(mentor_info_dict)
    return mentor_info_dict


def change_email_info_name(sheet, mentee_info_dict, mentor_info_dict):
    for row in sheet.rows:
        for col in row:
            if not isinstance(col.value, int):
                if col.value in mentee_info_dict.keys():
                    for info in mentee_info_dict[col.value][0:1]:
                        print("Mentee: {:30}".format(str(info)), end="")
                    print("{:30}".format(col.value + ","), end="")
            else:
                print("Mentor: {:50}".format(mentor_info_dict[col.value]), end=" ")
        print()


if __name__ == "__main__":
    mentor_mentee_index_sheet_path = "/Users/yuyang/Downloads/Mentor & Mentee序号.xlsx"
    sheet_info = read_07_Excel_sheet(mentor_mentee_index_sheet_path, "Mentee")
    # print_entire_sheet(sheet_info)
    mentee_info_dict = get_mentee_info(sheet_info)

    print()
    path = "/Users/yuyang/Downloads/Mentee选择Mentor意向和感谢信 - for all (Responses).xlsx"
    sheet = read_07_Excel_sheet(path, "Form Responses 1")
    mentor_mentee_dict_1st = get_selected_mentors(sheet, mentee_info_dict)
    print("1st round total chosen mentor numbers:", len(mentor_mentee_dict_1st))
    print("mentor_mentee_dict_1st", len(mentor_mentee_dict_1st), mentor_mentee_dict_1st)

    sheet_info = read_07_Excel_sheet(mentor_mentee_index_sheet_path, "Mentor")
    mentor_info_dict = get_mentor_info(sheet_info)

    print()
    sheet = read_07_Excel_sheet(path, "Statistics 1st round")
    print("Statistics 1st round")
    change_email_info_name(sheet, mentee_info_dict, mentor_info_dict)

    print()
    # two round overall statistics
    # sheet = read_07_Excel_sheet(path, "Statistics all")
    # print("Statistics all")
    # change_email_info_name(sheet, mentee_info_dict, mentor_info_dict)
